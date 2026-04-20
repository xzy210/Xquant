from __future__ import annotations

import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd


def normalize_row(values):
    return ["" if value is None else str(value) for value in values]


def _root_ns(root: ET.Element) -> str:
    if root.tag.startswith("{") and "}" in root.tag:
        return root.tag[1:].split("}", 1)[0]
    return ""


def _local_name(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _xlsx_xml_inspect(path: Path) -> dict:
    pkg_rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    workbook_sheets: list[tuple[str, str]] = []
    shared_strings: list[str] = []
    with zipfile.ZipFile(path) as zf:
        worksheet_entries = sorted(
            name for name in zf.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.iter():
                if _local_name(si.tag) != "si":
                    continue
                text = "".join((node.text or "") for node in si.iter() if _local_name(node.tag) == "t")
                shared_strings.append(text)
        try:
            workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
            rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            ns_main = {"main": _root_ns(workbook_root)}
            rel_ns = _root_ns(rel_root)
            ns_rel = {"rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
            rel_map = {
                rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
                for rel in rel_root.findall(f"{{{rel_ns}}}Relationship")
            }
            for sheet in workbook_root.findall("main:sheets/main:sheet", ns_main):
                name = sheet.attrib.get("name", "")
                rid = sheet.attrib.get(f"{{{ns_rel['rel']}}}id", "")
                workbook_sheets.append((name, rel_map.get(rid, "")))
        except Exception:
            workbook_sheets = []
        if not workbook_sheets:
            workbook_sheets = [
                (Path(name).stem, name.removeprefix("xl/"))
                for name in worksheet_entries
            ]

        result: dict[str, object] = {
            "path": str(path),
            "sheet_names": [name for name, _ in workbook_sheets],
            "sheets": {},
        }

        for index, (sheet_name, target) in enumerate(workbook_sheets):
            xml_path = target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"
            if xml_path not in zf.namelist() and index < len(worksheet_entries):
                xml_path = worksheet_entries[index]
            sheet_root = ET.fromstring(zf.read(xml_path))
            rows_out: list[list[str]] = []
            rows = [node for node in sheet_root.iter() if _local_name(node.tag) == "row"][:8]
            for row in rows:
                values: list[str] = []
                for cell in [node for node in row if _local_name(node.tag) == "c"]:
                    cell_type = cell.attrib.get("t", "")
                    value_node = next((node for node in cell.iter() if _local_name(node.tag) == "v"), None)
                    inline_node = next((node for node in cell.iter() if _local_name(node.tag) == "t"), None)
                    raw = ""
                    if inline_node is not None and inline_node.text is not None:
                        raw = inline_node.text
                    elif value_node is not None and value_node.text is not None:
                        raw = value_node.text
                    if cell_type == "s" and raw.isdigit():
                        idx = int(raw)
                        raw = shared_strings[idx] if 0 <= idx < len(shared_strings) else raw
                    values.append(raw)
                rows_out.append(values)
            non_empty_rows = [row for row in rows_out if any(str(cell).strip() for cell in row)]
            headers = non_empty_rows[0] if non_empty_rows else []
            preview = non_empty_rows[1:4] if len(non_empty_rows) > 1 else []
            result["sheets"][sheet_name] = {
                "max_row": len(rows_out),
                "max_column": max((len(row) for row in rows_out), default=0),
                "headers": headers,
                "preview_rows": preview,
                "raw_rows": non_empty_rows[:6],
            }
        return result


def inspect_file(path_str: str) -> dict:
    path = Path(path_str)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        try:
            sheet_names = list(pd.ExcelFile(path).sheet_names)
        except Exception:
            sheet_names = []
    if not sheet_names:
        workbook.close()
        return _xlsx_xml_inspect(path)
    result: dict[str, object] = {
        "path": str(path),
        "sheet_names": sheet_names,
        "sheets": {},
    }
    for sheet_name in sheet_names:
        try:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(min_row=1, max_row=8, values_only=True))
            normalized_rows = [normalize_row(row) for row in rows]
            non_empty_rows = [row for row in normalized_rows if any(cell.strip() for cell in row)]
            max_row = sheet.max_row
            max_column = sheet.max_column
        except Exception:
            df = pd.read_excel(path, sheet_name=sheet_name, nrows=8)
            non_empty_rows = [normalize_row(row) for row in df.itertuples(index=False, name=None)]
            non_empty_rows.insert(0, [str(col) for col in df.columns.tolist()])
            max_row = len(non_empty_rows)
            max_column = len(non_empty_rows[0]) if non_empty_rows else 0
        headers = non_empty_rows[0] if non_empty_rows else []
        preview = non_empty_rows[1:4] if len(non_empty_rows) > 1 else []
        result["sheets"][sheet_name] = {
            "max_row": max_row,
            "max_column": max_column,
            "headers": headers,
            "preview_rows": preview,
            "raw_rows": non_empty_rows[:6],
        }
    workbook.close()
    return result


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        raise SystemExit("usage: inspect_broker_history_xlsx.py [--output path] <file1> [file2 ...]")
    args = list(argv[1:])
    output_path: Path | None = None
    if len(args) >= 2 and args[0] == "--output":
        output_path = Path(args[1])
        args = args[2:]
    if not args:
        raise SystemExit("usage: inspect_broker_history_xlsx.py [--output path] <file1> [file2 ...]")
    payload = [inspect_file(arg) for arg in args]
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    if output_path is not None:
        output_path.write_text(text, encoding="utf-8")
        print("OK")
        return 0
    print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
